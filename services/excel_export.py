"""Excel export for report previews (openpyxl)."""

from __future__ import annotations

from datetime import datetime
from decimal import Decimal, InvalidOperation
from io import BytesIO
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

from core.i18n import t
from pages.settings import load_settings_values
from services.pdf_export import resolve_product_image_path

PROJECT_ROOT = Path(__file__).resolve().parent.parent
EXPORTS_DIR = PROJECT_ROOT / "exports"

_CURRENCY_FIELDS: frozenset[str] = frozenset(
    {
        "total_amount",
        "paid_amount",
        "remaining_amount",
        "total",
        "paid",
        "remaining",
        "unit_price",
        "inventory_value",
        "debt",
        "revenue",
        "total_spent",
        "total_revenue",
        "outstanding_debt",
        "warehouse_value",
        "current_price",
    }
)

_CURRENCY_NUMBER_FORMAT = '#,##0.00'


def _to_number(value: Any) -> float | None:
    if value is None or value == "":
        return None
    if isinstance(value, (int, float)):
        return float(value)
    try:
        return float(Decimal(str(value).strip().replace(",", "")))
    except (InvalidOperation, ValueError, AttributeError):
        return None


def _auto_column_width(worksheet: Any) -> None:
    for column_cells in worksheet.columns:
        max_length = 0
        column_letter = get_column_letter(column_cells[0].column)
        for cell in column_cells:
            value = "" if cell.value is None else str(cell.value)
            max_length = max(max_length, len(value))
        worksheet.column_dimensions[column_letter].width = min(max(max_length + 2, 12), 48)


def build_report_workbook(
    *,
    report_title: str,
    columns: list[dict[str, Any]],
    rows: list[dict[str, Any]],
    date_from: str | None = None,
    date_to: str | None = None,
) -> bytes:
    """Build a single-worksheet workbook for the currently displayed report."""
    settings = load_settings_values()
    company_name = settings.get("company_name") or "FlowCore"
    company_address = settings.get("company_address") or ""
    company_phone = settings.get("company_phone") or ""
    company_telegram = settings.get("company_telegram") or ""
    currency = settings.get("currency") or ""

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = t("reports.export.worksheet")

    bold = Font(bold=True)
    meta_font = Font(size=11)

    row_index = 1
    worksheet.cell(row=row_index, column=1, value=company_name).font = Font(bold=True, size=14)
    row_index += 1

    if company_address:
        worksheet.cell(row=row_index, column=1, value=company_address).font = meta_font
        row_index += 1
    contact_parts = []
    if company_phone:
        contact_parts.append(company_phone)
    if company_telegram:
        contact_parts.append(t("reports.export.telegram", telegram=company_telegram))
    if contact_parts:
        worksheet.cell(row=row_index, column=1, value=" | ".join(contact_parts)).font = meta_font
        row_index += 1
    if currency:
        worksheet.cell(
            row=row_index,
            column=1,
            value=t("reports.export.currency", currency=currency),
        ).font = meta_font
        row_index += 1

    row_index += 1
    worksheet.cell(row=row_index, column=1, value=report_title).font = Font(bold=True, size=12)
    row_index += 1

    generated_at = datetime.now().strftime("%Y-%m-%d %H:%M")
    generated_line = t("reports.export.generated", date=generated_at)
    if date_from or date_to:
        period = f"{date_from or '…'} — {date_to or '…'}"
        generated_line = (
            f"{generated_line} | {t('reports.export.period', period=period)}"
        )
    worksheet.cell(
        row=row_index,
        column=1,
        value=generated_line,
    ).font = meta_font
    row_index += 2

    export_columns = [
        column
        for column in columns
        if column.get("field") and column.get("field") not in {"status_code", "movement_type_code"}
    ]
    header_row = row_index
    for col_index, column in enumerate(export_columns, start=1):
        cell = worksheet.cell(
            row=header_row,
            column=col_index,
            value=str(column.get("label") or column.get("field") or ""),
        )
        cell.font = bold
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for data_offset, row in enumerate(rows):
        excel_row = header_row + 1 + data_offset
        for col_index, column in enumerate(export_columns, start=1):
            field = str(column.get("field") or "")
            raw_value = row.get(field, "")
            cell = worksheet.cell(row=excel_row, column=col_index)
            if field in _CURRENCY_FIELDS:
                numeric = _to_number(raw_value)
                if numeric is not None:
                    cell.value = numeric
                    cell.number_format = _CURRENCY_NUMBER_FORMAT
                    cell.alignment = Alignment(horizontal="right")
                else:
                    cell.value = raw_value
            elif isinstance(raw_value, (int, float)) and not isinstance(raw_value, bool):
                cell.value = raw_value
                cell.alignment = Alignment(horizontal="right")
            else:
                numeric = _to_number(raw_value) if field in {"quantity", "quantity_sold", "items_count"} else None
                if numeric is not None and str(raw_value).replace(".", "", 1).isdigit():
                    cell.value = int(numeric) if float(numeric).is_integer() else numeric
                    cell.alignment = Alignment(horizontal="right")
                else:
                    cell.value = raw_value if raw_value is not None else ""

    last_col = max(len(export_columns), 1)
    last_data_row = header_row + max(len(rows), 1)
    worksheet.auto_filter.ref = (
        f"A{header_row}:{get_column_letter(last_col)}{last_data_row}"
    )
    worksheet.freeze_panes = f"A{header_row + 1}"
    _auto_column_width(worksheet)

    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def save_report_workbook(
    content: bytes,
    *,
    report_type: str,
) -> Path:
    """Persist workbook under exports/ and return the path."""
    EXPORTS_DIR.mkdir(parents=True, exist_ok=True)
    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    safe_type = (report_type or "report").strip().lower().replace(" ", "_") or "report"
    path = EXPORTS_DIR / f"{safe_type}_{stamp}.xlsx"
    path.write_bytes(content)
    return path


def build_catalog_workbook(
    *,
    report_title: str,
    columns: list[dict[str, Any]],
    rows: list[dict[str, Any]],
) -> bytes:
    """Build a product catalog workbook (supports optional image column)."""
    settings = load_settings_values()
    company_name = settings.get("company_name") or "FlowCore"
    company_address = settings.get("company_address") or ""
    company_phone = settings.get("company_phone") or ""
    company_telegram = settings.get("company_telegram") or ""
    currency = settings.get("currency") or ""

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = t("reports.export.worksheet")

    bold = Font(bold=True)
    meta_font = Font(size=11)

    row_index = 1
    worksheet.cell(row=row_index, column=1, value=company_name).font = Font(bold=True, size=14)
    row_index += 1

    if company_address:
        worksheet.cell(row=row_index, column=1, value=company_address).font = meta_font
        row_index += 1
    contact_parts = []
    if company_phone:
        contact_parts.append(company_phone)
    if company_telegram:
        contact_parts.append(t("reports.export.telegram", telegram=company_telegram))
    if contact_parts:
        worksheet.cell(row=row_index, column=1, value=" | ".join(contact_parts)).font = meta_font
        row_index += 1
    if currency:
        worksheet.cell(
            row=row_index,
            column=1,
            value=t("reports.export.currency", currency=currency),
        ).font = meta_font
        row_index += 1

    row_index += 1
    worksheet.cell(row=row_index, column=1, value=report_title).font = Font(bold=True, size=12)
    row_index += 1
    generated_at = datetime.now().strftime("%Y-%m-%d %H:%M")
    worksheet.cell(
        row=row_index,
        column=1,
        value=t("reports.export.generated", date=generated_at),
    ).font = meta_font
    row_index += 2

    export_columns = [
        column
        for column in columns
        if column.get("field") and column.get("field") not in {"status_code", "movement_type_code"}
    ]
    header_row = row_index
    image_col_index: int | None = None
    for col_index, column in enumerate(export_columns, start=1):
        field = str(column.get("field") or "")
        if field in {"image_url", "image", "product_image"}:
            image_col_index = col_index
        cell = worksheet.cell(
            row=header_row,
            column=col_index,
            value=str(column.get("label") or column.get("field") or ""),
        )
        cell.font = bold
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for data_offset, row in enumerate(rows):
        excel_row = header_row + 1 + data_offset
        if image_col_index is not None:
            worksheet.row_dimensions[excel_row].height = 48
        for col_index, column in enumerate(export_columns, start=1):
            field = str(column.get("field") or "")
            cell = worksheet.cell(row=excel_row, column=col_index)
            if field in {"image_url", "image", "product_image"}:
                image_path = resolve_product_image_path(row.get("image_url") or row.get(field))
                if image_path.exists():
                    try:
                        xl_image = XLImage(str(image_path))
                        xl_image.width = 54
                        xl_image.height = 54
                        worksheet.add_image(xl_image, f"{get_column_letter(col_index)}{excel_row}")
                    except Exception:
                        cell.value = ""
                else:
                    cell.value = ""
                continue
            raw_value = row.get(field, "")
            if field in _CURRENCY_FIELDS:
                numeric = _to_number(raw_value)
                if numeric is not None:
                    cell.value = numeric
                    cell.number_format = _CURRENCY_NUMBER_FORMAT
                    cell.alignment = Alignment(horizontal="right")
                else:
                    cell.value = raw_value
            elif isinstance(raw_value, (int, float)) and not isinstance(raw_value, bool):
                cell.value = raw_value
                cell.alignment = Alignment(horizontal="right")
            else:
                cell.value = raw_value if raw_value is not None else ""

    last_col = max(len(export_columns), 1)
    last_data_row = header_row + max(len(rows), 1)
    worksheet.auto_filter.ref = (
        f"A{header_row}:{get_column_letter(last_col)}{last_data_row}"
    )
    worksheet.freeze_panes = f"A{header_row + 1}"
    if image_col_index is not None:
        worksheet.column_dimensions[get_column_letter(image_col_index)].width = 12
    _auto_column_width(worksheet)

    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()
